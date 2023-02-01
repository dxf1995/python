
from openpyxl import load_workbook

workbook = load_workbook(filename = "D:/learn/python/day11/CPU详细参数汇总.xlsx")

#获取sheet名称
print(workbook.sheetnames)

sheet = workbook["志强3代Icelake"]
print(sheet)